import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import time

# Streamlit UI
st.title("Automated Report Generator")

# User Inputs
report_date_from = st.date_input("Report Date From")
report_date_to = st.date_input("Report Date To")

# Configurations
robot_path = r'\\nseura0090\DXCITO\NBSES\DMT\02 TOOLS\Nestle_EDF DE'
smtp_user = 'justyna.brzezniakiewicz@nestle.com'
smtp_password = 'Zadanie_dozrobienia2024'

# Read Config File
config_path = os.path.join(robot_path, 'data', 'Config.xlsx')
wb = openpyxl.load_workbook(config_path, data_only=True)
ws = wb['Config']

payroll_folder = ws.cell(row=4, column=4).value
email_to = ws.cell(row=2, column=1).value
email_account = ws.cell(row=2, column=2).value
signature = ws.cell(row=2, column=3).value
email_folder = ws.cell(row=4, column=1).value
folder = ws.cell(row=4, column=2).value
rm = ws.cell(row=4, column=3).value

ws = wb['Emails_Recon']
emails_recon = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]

ws = wb['Request_types']
req_codes = [[ws.cell(row=i, column=j).value for j in range(1, 8)] for i in range(1, ws.max_row + 1)]

ws = wb['Out of scope']
out_of_scope = [[ws.cell(row=i, column=j).value for j in range(1, 3)] for i in range(1, ws.max_row + 1)]

# Format Dates
formatted_date_from = report_date_from.strftime("%m.%d.%Y")
formatted_date_to = report_date_to.strftime("%m.%d.%Y")
formatted_date_from_report_name = formatted_date_from.replace(".", "/")
formatted_date_to_report_name = formatted_date_to.replace(".", "/")

# Create Subfolder
current_date_txt = datetime.now().strftime("%m.%d.%Y")
current_subfolder = os.path.join(folder, current_date_txt)
os.makedirs(current_subfolder, exist_ok=True)

# Web Automation
options = webdriver.ChromeOptions()
options.add_argument('--headless')
driver = webdriver.Chrome(options=options)
driver.get(rm)

# Handle potential login
try:
    account_element = driver.find_element(By.XPATH, "//div[contains(text(), 'Pick an account')]")
    account_element.click()
    time.sleep(4)
except:
    pass

# Set dropdown values and populate text fields
Select(driver.find_element(By.ID, "phlMain_ddlRegion")).select_by_visible_text("Germany")
Select(driver.find_element(By.ID, "phlMain_ddlCountry")).select_by_visible_text("Germany")
Select(driver.find_element(By.ID, "phlMain_ddlReferenceDate")).select_by_visible_text("Closing")
Select(driver.find_element(By.ID, "phlMain_ddlStatus")).select_by_visible_text("Closed")

date_from_element = driver.find_element(By.ID, "x:362216426.0:mkr:3")
date_from_element.send_keys(Keys.CONTROL + "a")
date_from_element.send_keys(formatted_date_from)

date_to_element = driver.find_element(By.ID, "x:362216426.0:mkr:3")
date_to_element.send_keys(Keys.CONTROL + "a")
date_to_element.send_keys(formatted_date_to)

# Export report
driver.find_element(By.ID, "phlMain_btnExport").click()
time.sleep(3)

# Save report
os.makedirs(os.path.join(robot_path, 'reports'), exist_ok=True)
report_path = os.path.join(robot_path, 'reports', 'EDF Report_' + formatted_date_from_report_name + ' - ' + formatted_date_to_report_name + '.xlsx')
time.sleep(5)
driver.close()

# Move the report file
os.rename(report_path, os.path.join(robot_path, 'reports', 'EDF Report_' + formatted_date_from_report_name + ' - ' + formatted_date_to_report_name + '.xlsx'))

# Prepare the report
prepare_report_path = os.path.join(robot_path, 'data', 'PrepareReport.xlsm')
wb_macro = openpyxl.load_workbook(prepare_report_path, keep_vba=True)
ws_macro = wb_macro['Path']
ws_macro.cell(row=1, column=1).value = report_path
wb_macro.save(prepare_report_path)

# Close the Excel file
wb_macro.close()

st.success("Report generated and saved successfully!")

# # Send email notification
# msg = MIMEMultipart()
# msg['From'] = smtp_user
# msg['To'] = email_to
# msg['Subject'] = 'Automated Report'
# body = 'Please find the attached report.'
# msg.attach(MIMEText(body, 'plain'))
#
# with open(report_path, "rb") as attachment:
#     part = MIMEBase("application", "octet-stream")
#     part.set_payload(attachment.read())
#     encoders.encode_base64(part)
#     part.add_header(
#         "Content-Disposition",
#         f"attachment; filename= {os.path.basename(report_path)}",
#     )
#     msg.attach(part)
#
# server = smtplib.SMTP('smtp.gmail.com', 587)
# server.starttls()
# server.login(smtp_user, smtp_password)
# text = msg.as_string()
# server.sendmail(smtp_user, email_to, text)
# server.quit()

st.success("Email sent successfully!")
